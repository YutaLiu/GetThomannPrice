#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Thu Jan 30 20:40:38 2020

@author: liuyuda
"""

import openpyxl
from bs4 import BeautifulSoup
import requests
import re
#wb = Workbook()

def write_xlsm():

    wb = openpyxl.load_workbook(filename='Shopee_mass_upload_template_sku_tw.xlsm', read_only=False, keep_vba=True)
    
    print(wb.sheetnames)
    
    sheet = wb.active
    
    start_col = 7
    start_row = ord('A')
    sheet['A5'].value
    value = sheet['A5']



def parsing_thomman():
    return 0


usb_interface = 'https://www.thomann.de/gb/usb_audio_interfaces.html'

res = requests.get(usb_interface, timeout=30)
#print(res.text)
soup = BeautifulSoup(res.text, 'html.parser')
soup.encoding = 'utf-8' 
#soup = BeautifulSoup(res.text,'lxml')
# =============================================================================
# a_tags = soup.find_all('div')
# for tag in a_tags:
#     
#     if tag != None:
#         print(tag.string , tag == None)
# =============================================================================

model = []
price = []
for item in soup.body.find_all('div', {'class':'name'}):
    model.append(item.string)


for item in soup.body.find_all('span', {'class':'primary bold'}):
    try:
        value = int(re.findall(r'\b\d+\b', item.string)[1]) + int(re.findall(r'\b\d+\b', item.string)[0])*1000 
    except:
        value = int(re.findall(r'\b\d+\b', item.string)[0]) 
    price.append(value*40)
    
result = list(zip(model, price))

for i,j in result:
    print ("item= ",i,"NTD Price=",j)